import json
from typing import Any

from sqlalchemy.orm import Session

from app.engine.decision.decision_logger import update_decision_engine_outcome
from app.engine.setup.setup_logger import update_setup_outcome
from app.engine.specialist.models import LabelRecord, SpecialistEngineLog
from app.engine.specialist.shadow_logger import update_market_result

EXPECTED_FIELDS = [
    "date",
    "time",
    "signal_id",
    "evaluation_id",
    "instrument",
    "setup",
    "action_taken",
    "direction",
    "strike",
    "entry_price",
    "exit_price",
    "result",
    "r_multiple",
    "confidence",
    "wrong_block_reason",
    "notes",
]


def _safe_float(value: Any) -> float | None:
    try:
        return float(value) if value not in (None, "", "null") else None
    except Exception:
        return None


def _safe_int(value: Any) -> int | None:
    try:
        return int(float(value)) if value not in (None, "", "null") else None
    except Exception:
        return None


def _direction(value: Any, fallback: Any = None) -> str | None:
    raw = str(value or fallback or "").upper()
    if "PUT" in raw or raw == "PE":
        return "PE"
    if "CALL" in raw or raw == "CE":
        return "CE"
    return raw or None


def _result(row: dict) -> str | None:
    value = row.get("result")
    if value:
        return str(value).upper()
    pnl = _safe_float(row.get("pnl_points") or row.get("pnl"))
    if pnl is None:
        return None
    if pnl > 0:
        return "WIN"
    if pnl < 0:
        return "LOSS"
    return "BREAKEVEN"


def _row_to_label(row: dict) -> dict:
    row = {str(k).strip().lower(): v for k, v in row.items()}
    return {
        "date_str": str(row.get("date", "") or ""),
        "time_str": str(row.get("time") or row.get("entry_time") or ""),
        "signal_id": row.get("signal_id") or None,
        "evaluation_id": row.get("evaluation_id") or None,
        "instrument": row.get("instrument") or row.get("underlying") or None,
        "setup": row.get("setup") or row.get("pattern") or None,
        "action_taken": row.get("action_taken") or row.get("label") or None,
        "direction": _direction(row.get("direction"), row.get("option_type")),
        "strike": _safe_int(row.get("strike")),
        "entry_price": _safe_float(row.get("entry_price") or row.get("entry_premium")),
        "exit_price": _safe_float(row.get("exit_price") or row.get("exit_premium")),
        "result": _result(row),
        "r_multiple": _safe_float(row.get("r_multiple")),
        "confidence": _safe_float(row.get("confidence")),
        "wrong_block_reason": row.get("wrong_block_reason") or None,
        "notes": row.get("notes") or row.get("why_taken") or row.get("reason") or None,
        "label_source": "MANUAL",
    }


def _market_result(data: dict) -> str | None:
    direction = (data.get("direction") or "").upper()
    result = (data.get("result") or "").upper()
    action = (data.get("action_taken") or "").upper()
    if direction in ("CE", "PE") and result in ("WIN", "LOSS", "BREAKEVEN"):
        return f"{direction}_{result}"
    if action in ("NO_TRADE_CORRECT", "NO_TRADE_MISSED"):
        return action
    if action == "MISSED":
        return "NO_TRADE_MISSED"
    if action == "NO_TRADE_AVOID":
        return "NO_TRADE_CORRECT"
    return None


def _process_rows(db: Session, rows: list) -> dict:
    total = len(rows)
    matched = 0
    standalone = 0
    errors = []
    for index, row in enumerate(rows):
        try:
            data = _row_to_label(row)
            label = LabelRecord(**data)
            db.add(label)
            eval_id = data.get("evaluation_id")
            market_result = _market_result(data)
            if eval_id and market_result:
                matched_engines = _propagate_market_result(
                    db=db,
                    evaluation_id=eval_id,
                    market_result=market_result,
                )
                if matched_engines > 0:
                    matched += 1
                else:
                    standalone += 1
            else:
                standalone += 1
        except Exception as e:
            errors.append(f"Row {index}: {e}")
    db.commit()
    return {
        "total_records": total,
        "matched_to_engine_logs": matched,
        "stored_as_standalone": standalone,
        "errors": errors,
    }


def _propagate_market_result(db: Session, evaluation_id: str, market_result: str) -> int:
    engine_names = [
        row[0]
        for row in (
            db.query(SpecialistEngineLog.engine_name)
            .filter(SpecialistEngineLog.evaluation_id == evaluation_id)
            .distinct()
            .all()
        )
    ]
    for engine_name in engine_names:
        update_market_result(db, evaluation_id, engine_name, market_result, "MANUAL")
    if engine_names:
        update_setup_outcome(db, evaluation_id, market_result)
        update_decision_engine_outcome(db, evaluation_id, market_result)
    return len(engine_names)


def import_jsonl_labels(db: Session, file_path: str) -> dict:
    rows = []
    errors = []
    with open(file_path, "r", encoding="utf-8") as handle:
        for line_no, line in enumerate(handle, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except Exception as e:
                errors.append(f"Line {line_no}: {e}")
    result = _process_rows(db, rows)
    result["errors"].extend(errors)
    return result


def import_xlsx_labels(db: Session, file_path: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {
            "total_records": 0,
            "matched_to_engine_logs": 0,
            "stored_as_standalone": 0,
            "errors": ["openpyxl not installed. Run: pip install openpyxl"],
        }
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    try:
        worksheet = workbook.active
        headers = [str(cell.value or "").strip().lower() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        rows = []
        for worksheet_row in worksheet.iter_rows(min_row=2, values_only=True):
            rows.append({headers[index]: value for index, value in enumerate(worksheet_row) if index < len(headers)})
        return _process_rows(db, rows)
    finally:
        workbook.close()
